from telegram import Update
from telegram.ext import ContextTypes
from asyncio import sleep

import openpyxl
from aiohttp import ClientSession
from json import dumps as json_dumps

import logging
from logging.handlers import RotatingFileHandler

import re
from random import choices as random_choices

from keyboards import (
    start_keyboard,
    account_keyboard, account_add_balance,
    back_keyboard,
    confirm_add_keyboard, confirm_invoice_keyboard, pays_keyboard)
from database import (insert_tasks_db, insert_users_db, update_users_db, select_users_db,
                      update_time_weight_links, select_tasks, insert_links_db, select_links,
                      insert_link_transitions_db, select_pays, insert_pays)

from specs import specs, states

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)
logger.addHandler(RotatingFileHandler(filename=f"{specs.logs_path}{__name__}.log",
                                      mode='w',
                                      maxBytes=1024 * 1024))


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /start is issued."""
    try:
        user_id = update.message.from_user.id
        await insert_users_db(user_id=user_id, balance_hl=10, balance=5000)
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text='Привет, я бот для ссылок!',
            reply_markup=start_keyboard
        )
        await set_checkout(update=update,context=context)
        logger.info(msg=f'Succeed to start')

        return states.START
    except Exception as e:
        logger.exception(msg=f'Failed to start: {e}')


async def task_complete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get screenshot of completed task and add points"""

    user_id = update.message.from_user.id

    photo_id = update.message.photo[-1].file_id
    try:
        new_photo = await context.bot.get_file(file_id=photo_id)
        await new_photo.download_to_drive(f'images/{photo_id}.jpg')
    except Exception as e:
        logger.exception(msg=f'Failed to download photo: {e}')

    try:
        _, _, task, task_id = await select_users_db(user_id=user_id, column=-1)
        if task:

            await update_users_db(user_id=user_id, balance_hl=1, task='')
            balance_hl = await select_users_db(user_id=user_id, column=1)
            await insert_tasks_db(user_id=user_id, task=task, photo_id=photo_id, task_id=task_id)
            count_pays, sum_pays = await select_pays(user_id=user_id)
            if not count_pays:
                count_pays = [0]
                sum_pays = [0]
            try:
                wb = openpyxl.load_workbook(f'excel/db.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            ws = wb.worksheets[0]
            index = ws.max_row + 1
            ws.cell(row=index, column=1, value=f'{user_id}')
            ws.cell(row=index, column=2, value=f'{task}')
            ws.cell(row=index, column=3, value=f'{photo_id}')

            ws.cell(row=index, column=4, value=f'{count_pays[0]}')
            ws.cell(row=index, column=5, value=f'{sum_pays[0] * 0.01}')
            wb.save('excel/db.xlsx')
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'{update.message.from_user.username} получили 1 ХЛБалл! Ваш баланс: {balance_hl}',
                reply_markup=back_keyboard
            )

            logger.info(msg=f'Succeed to task complete 0')
            return states.GET_LINK

        else:
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'У тебя нет задания',
                reply_markup=start_keyboard
            )
            logger.warning(msg=f'Rejected to task complete by empty task')
            return states.START
    except Exception as e:
        logger.exception(msg=f'Failed to task complete: {e}')


async def get_link(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        await update_time_weight_links()
        user_id = update.message.from_user.id
        if await select_users_db(user_id=user_id, column=2):
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'Сначала выполни текущее задание',
                reply_markup=back_keyboard
            )
            return states.GET_LINK

        rows = await select_tasks(user_id=user_id)
        links = []
        weights = []
        if rows:
            for row in rows:
                links.append((row[0], row[2]))
                weights.append(row[1] + 1)

        else:
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'Пока что нет задач',
                reply_markup=start_keyboard
            )
            return states.START

        task, task_id = random_choices(population=links, weights=weights)[0]

        await update_users_db(user_id=user_id, task=task, task_id=task_id)
        # link_id = await select_links(user_id=user_id, link=task)
        # link_id = link_id[0]
        await insert_link_transitions_db(link_id=task_id)
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text=f'Отправь скриншот выполненной задачи. Задача: {task}',
            reply_markup=back_keyboard
        )

        logger.info(msg=f'Succeed to get link')
        return states.GET_LINK
    except Exception as e:
        logger.exception(msg=f'Failed to get link: {e}')


async def send_link(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text=f'Выбери оплату',
            reply_markup=confirm_add_keyboard
        )

        logger.info(msg=f'Succeed to send link')
        return states.SEND_LINK
    except Exception as e:
        logger.exception(msg=f'Failed to send link: {e}')


async def confirm_add(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        message = update.message.text
        user_id = update.message.from_user.id
        if message == 'Добавить за 50 рублей и 10 ХЛБаллов':
            specs.choose_cost.update({user_id: 1})
        else:
            specs.choose_cost.update({user_id: 2})

        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text=f'Отправь ссылку с небольшим описанием того, что нужно сделать в форме: <<<тут ссылка и описание>>>',
            reply_markup=back_keyboard
        )

        logger.info(msg=f'Succeed to comfirm_add')
        return states.ACCEPT_LINK
    except Exception as e:
        logger.exception(msg=f'Failed to comfirm_add: {e}')


async def add_link(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    pattern = r'((https?://)[\s\S]*?)'
    message = update.message.text

    match = re.search(pattern, message)

    if not match:
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text=f'Ссылка не найдена'
        )
        logger.info(msg=f'Rejected to add link by link not found')
        return states.ACCEPT_LINK

    try:
        user_id = update.message.from_user.id
        balance, balance_hl, _, _ = await select_users_db(user_id=user_id, column=-1)
        if specs.choose_cost.get(user_id) == 1 and balance >= specs.price[0] and balance_hl >= specs.price_hl[0]:
            await update_users_db(user_id=user_id, balance=-specs.price[0], balance_hl=-specs.price_hl[0])
            # _start, _end = match.span()
            link = f'{message}'
            await insert_links_db(user_id=user_id, link=link)

            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'Ссылка добавлена',
                reply_markup=start_keyboard
            )
            logger.info(msg=f'Succeed to add link')
            return states.START
        elif specs.choose_cost.get(user_id) == 2 and balance >= specs.price[1] and balance_hl >= specs.price_hl[1]:
            await update_users_db(user_id=user_id, balance=-specs.price[1], balance_hl=-specs.price_hl[1])
            _start, _end = match.span()
            link = f'{message[_start:_end]}'
            await insert_links_db(user_id=user_id, link=link)

            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'Ссылка добавлена',
                reply_markup=start_keyboard
            )
            logger.info(msg=f'Succeed to add link')
            return states.START

        else:
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=f'Недостаточно денег на счёте. Баланс: {balance / 100}',
                reply_markup=start_keyboard
            )
            logger.info(msg=f'Rejected to add link by low balance')
            return states.START

    except Exception as e:
        logger.exception(msg=f'Failed to add link: {e}')


async def personal_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id
    try:
        balance, balance_hl, task, _ = await select_users_db(user_id=user_id, column=-1)

        if task is None:
            task = 'нет'
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text=f'Ваш id: {user_id}\n'
                 f'Баланс: у вас {balance // 100}.{balance % 100} рублей\n'
                 f'ХЛ: у вас {balance_hl} ХЛБаллов\n'
                 f'Задание для выполнения: {task}\n',
            reply_markup=account_keyboard
        )
        links = await select_links(user_id=user_id)
        if links:
            answer = f'{links[0][0]}'
            for link in links[1::]:
                answer = f'{answer} ||| {link[0]}'
            await update.message.reply_text(
                reply_to_message_id=update.message.message_id,
                text=answer,
                reply_markup=account_keyboard
            )
        logger.info(msg=f'Succeed check personal account')
        return states.ACCOUNT
    except Exception as e:
        logger.exception(msg=f'Failed to check personal account: {e}')


async def back(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.message.from_user.id
        specs.choose_cost.pop(user_id, None)
        specs.payment_payload.pop(user_id, None)

        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text='Вернулись в меню',
            reply_markup=start_keyboard
        )

        logger.info(msg=f'Succeed to back')
        return states.START
    except Exception as e:
        logger.exception(msg=f'Failed to back: {e}')


async def account_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await update.message.reply_text(
            reply_to_message_id=update.message.message_id,
            text='Выбери платёж',
            reply_markup=account_add_balance
        )
        logger.info(msg=f'Succeed to account_payment')
        return states.ACCOUNT_ADD_BALANCE
    except Exception as e:
        logger.exception(msg=f'Failed to account_payment: {e}')


async def checkout(context: ContextTypes.DEFAULT_TYPE):
    """Check invoice success after 24 hours"""

    user_id = context.job.chat_id
    headers = {'accept': 'application/json',
               'X-Api-Key': specs.payment_token,
               }
    offerId = specs.payment_payload.get(user_id)
    for _ in range(5):
        continue_check = False
        try:
            async with ClientSession() as session:
                async with session.get(url=f'https://gate.lava.top/api/v1/invoices/{offerId}',
                                       headers=headers
                                       ) as response:
                    if response.status == 200:
                        data = await response.json()
                        if data['status'] == 'COMPLETED':
                            await context.bot.send_message(
                                text='Баланс успешно пополнен',
                                reply_markup=start_keyboard
                            )
                            await update_users_db(user_id=user_id, balance=int(data['receipt']['amount'] * 100))
                            await insert_pays(user_id=user_id, pays_sum=int(data['receipt']['amount'] * 100))
                            specs.payment_payload.pop(user_id, None)
                            logger.info(msg=f'Succeed to account_invoice_confirm')
                            return states.START
                        elif data['status'] == 'IN_PROGRESS':
                            await context.bot.send_message(
                                text='Транзакция выполняется',
                                reply_markup=start_keyboard
                            )
                            logger.warning(msg='Waited to account_invoice_confirm: in-progress')
                            return states.START
                        else:
                            print(data["status"])
                            error_codes = {'cancel': 'Транзакция отменена',
                                           'error': 'Во время транзакции возникла ошибка'}
                            await context.bot.send_message(
                                text='Какая-то ошибка',
                                reply_markup=start_keyboard
                            )
                            specs.payment_payload.pop(user_id, None)
                            logger.warning(msg=f'Failed to account_invoice_confirm: {data["status"]}')
                            return states.START

                    else:
                        await context.bot.send_message(
                            text='Что-то пошло не так',
                            reply_markup=start_keyboard
                        )
                        continue_check = True
                        return states.START

        except Exception as e:
            logger.exception(msg=f'Failed to checkout: {e}')
        if continue_check:
            pass
        else:
            break
        await sleep(3)



async def set_checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Add a job to the queue."""
    try:
        user_id = update.message.from_user.id
        context.job_queue.run_once(callback=checkout,
                                   when=60*60*24 + 5,
                                   chat_id=user_id,
                                   name=str(user_id),
                                   )

        logger.info(msg=f'Succeed to set_checkout')
    except Exception as e:
        logger.exception(msg=f'Failed to set_checkout: {e}')


async def account_send_invoice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send an invoice"""
    try:
        user_id = update.message.from_user.id
        # 4 products with different prices and IDs
        amount = {account_add_balance.keyboard[0][0].text: "094a1456-43e4-4bc1-8d17-60a44ea4c5ba",
                  account_add_balance.keyboard[1][0].text: "e7b0a03b-e980-4207-8b5f-1e9960abf084",
                  account_add_balance.keyboard[2][0].text: '139dc677-4314-4ede-99f0-176dfcf5fc18',
                  account_add_balance.keyboard[3][0].text: "e9e0ed91-c72b-4a78-bf80-70182d73f69f"
                  }
        headers = {'accept': 'application/json',
                   'X-Api-Key': specs.payment_token,
                   'Content-Type': 'application/json',
                   }

        data = {"email": "client@gmail.com",
                "offerId": amount[update.message.text],
                "currency": "RUB"
                }
        async with ClientSession() as session:
            async with session.post(url='https://gate.lava.top/api/v2/invoice',
                                    headers=headers,
                                    data=json_dumps(data)) as response:
                if response.status == 201:
                    data = await response.json()
                    if data['status'] == "in-progress":

                        specs.payment_payload.update({user_id: data['id']})
                        await update.message.reply_text(
                            text=f"Для оплаты нажмите на «Оплатить»",
                            reply_to_message_id=update.message.message_id,
                            reply_markup=pays_keyboard(url=data['paymentUrl'])
                        )
                        await update.message.reply_text(
                            text=f"После оплаты нажмите на «Оплачено»\n"
                                 f"Для возврата в главное меню нажмите «Назад»",
                            reply_to_message_id=update.message.message_id,
                            reply_markup=confirm_invoice_keyboard
                        )
                        logger.info(msg=f'Succeed to account_send_invoice:')
                        return states.ACCOUNT_CONFIRM_ADD
                    else:

                        await update.message.reply_text(
                            text='Что-то пошло не так, попробуй ещё раз',
                            reply_to_message_id=update.message.message_id,
                            reply_markup=start_keyboard
                        )
                        logger.error(msg=f'Failed to account_send_invoice.\n'
                                         f'Status: {response.status}\n'
                                         f'ErrorCode: {data["code"]}'
                                         f'Error: {data["message"]}')
                        return states.START
                else:
                    await update.message.reply_text(
                        text='Что-то пошло не так, попробуй ещё раз',
                        reply_to_message_id=update.message.message_id,
                        reply_markup=start_keyboard
                    )
                    logger.error(msg=f'Failed to account_send_invoice.\n'
                                     f'Status: {response.status}\n'
                                 )
                    return states.START

    except Exception as e:
        logger.exception(msg=f'Failed to account_send_invoice: {e}')


async def account_invoice_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:

        user_id = update.message.from_user.id
        headers = {'accept': 'application/json',
                   'X-Api-Key': specs.payment_token,
                   }
        # data = {'id': specs.payment_payload.get(user_id)}
        offerId = specs.payment_payload.get(user_id)
        async with ClientSession() as session:
            async with session.get(url=f'https://gate.lava.top/api/v1/invoices/{offerId}',
                                    headers=headers
                                    ) as response:
                if response.status == 200:
                    data = await response.json()
                    if data['status'] == "COMPLETED":
                        await update.message.reply_text(
                            text='Баланс успешно пополнен',
                            reply_to_message_id=update.message.message_id,
                            reply_markup=start_keyboard
                        )
                        await update_users_db(user_id=user_id, balance=int(data['receipt']['amount'] * 100))
                        await insert_pays(user_id=user_id, pays_sum=int(data['receipt']['amount'] * 100))
                        specs.payment_payload.pop(user_id, None)
                        logger.info(msg=f'Succeed to account_invoice_confirm')
                        return states.START
                    elif data['status'] == 'IN_PROGRESS':
                        await update.message.reply_text(
                            text='Оплата ещё не прошла. Транзакция выполняется',
                            reply_to_message_id=update.message.message_id,
                            reply_markup=confirm_invoice_keyboard
                        )
                        logger.info(msg='Waited to account_invoice_confirm: IN_PROGRESS')
                        return states.ACCOUNT_CONFIRM_ADD
                    else:
                        print(data["status"])
                        print(data["status"])
                        await update.message.reply_text(
                            text='Какая-то ошибка',
                            reply_to_message_id=update.message.message_id,
                            reply_markup=start_keyboard
                        )
                        specs.payment_payload.pop(user_id, None)
                        logger.warning(msg=f'Failed to account_invoice_confirm: {data["status"]} ')
                        return states.START

                else:
                    await update.message.reply_text(
                        text='Что-то пошло не так',
                        reply_to_message_id=update.message.message_id,
                        reply_markup=start_keyboard
                    )
                    return states.START

    except Exception as e:
        logger.exception(msg=f'Failed to account_invoice_confirm: {e}')


if __name__ == '__main__':
    pass
